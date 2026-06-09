"""
BIT Breakdown Table Acceptance Criteria Checker

Steps:
  1. Fetch available_balance + currency from Matrixport Wallet Balance API
  2. Fetch available_balance + currency from Matrixport Balance+ API
  3. Fetch token prices (USDT, USDC, TRX) — Matrixport RFQ + CoinGecko 24h
  4. Export Steps 1-3 to Excel
  5. Log in to DAM
  6. Open Portfolio by portfolioId
  7. Locate "David BIT Account" table on Overview page
  8. Export portfolio table to Excel (Token, Price, Price 24H, Total Amount, Value, Total Value)
  9. Compare: available_balance ↔ Total Amount, price ↔ Price (1% tolerance)
  10. Verify: Value = amount × price; Total Value = sum of values

Usage:
    python bit_breakdown_table_check.py --portfolio-id <ID>
    python bit_breakdown_table_check.py --portfolio-id <ID> --headed
    python bit_breakdown_table_check.py --portfolio-id <ID> --out report.xlsx
"""

import argparse
import hashlib
import hmac
import json
import time
from pathlib import Path

import requests

# ─── Credentials for "David BIT Account" ─────────────────────────────────
API_KEY = "ak-9c60c430-a9d5-483f-8d90-32494ab20022"
SECRET  = "qGtXrDOMsHHVca5dg4Rimcy0gqjz0mBCW3cs3sV1iRzYAmv6GQ1Pjdw1mgEu0vEU"

MATRIXPORT_BASE = "https://mapi.matrixport.com"
COINGECKO_BASE  = "https://api.coingecko.com/api/v3"

DAM_URL      = "https://dam-sit.mqbc21.com"
DAM_EMAIL    = "roninx688@gmail.com"
DAM_PASSWORD = "0987654321a@A"

# Tokens to price-check (Step 3)
TOKENS = ["USDT", "USDC", "TRX"]

COINGECKO_IDS = {
    "USDT": "tether",
    "USDC": "usd-coin",
    "TRX":  "tron",
}

# (Matrixport symbol, min_qty) — USDT/USDT = 1.0 by definition
MATRIXPORT_PAIRS = {
    "USDC": ("USDCUSDT", "20"),
    "TRX":  ("TRXUSDT",  "300"),
}

VERBOSE = False


def log(msg: str):
    print(msg)


def vlog(msg: str):
    if VERBOSE:
        print(f"  [verbose] {msg}")


# ─── Signing ──────────────────────────────────────────────────────────────────
def sign_v2(method: str, path: str, query_string: str) -> dict:
    ts = str(int(time.time() * 1000))
    prehash = ts + method + path + "&" + query_string
    sig = hmac.new(SECRET.encode(), prehash.encode(), hashlib.sha256).hexdigest()
    return {
        "X-MatrixPort-Access-Key": API_KEY,
        "X-Signature": sig,
        "X-Timestamp": ts,
        "X-Auth-Version": "v2",
    }


def mp_get(path: str, params: dict = None) -> dict:
    qs = ""
    if params:
        qs = "&".join(f"{k}={v}" for k, v in sorted(params.items()))
    headers = sign_v2("GET", path, qs)
    url = MATRIXPORT_BASE + path + (f"?{qs}" if qs else "")
    vlog(f"GET {url}")
    resp = requests.get(url, headers=headers, timeout=15)
    resp.raise_for_status()
    data = resp.json()
    vlog(f"Response (truncated): {json.dumps(data, indent=2)[:1000]}...")
    return data


# ─── Step 1: Wallet Balance ───────────────────────────────────────────────────
def fetch_wallet_balance() -> list[dict]:
    data = mp_get("/mapi/v1/wallet/balance")
    if data.get("code") != 0:
        raise RuntimeError(f"wallet/balance error: code={data.get('code')} msg={data.get('message')}")
    inner = data.get("data") or {}
    if isinstance(inner, list):
        records = inner
    else:
        records = (inner.get("items") or inner.get("assets") or
                   inner.get("list") or inner.get("balances") or [])
    result = []
    for r in records:
        cur   = r.get("currency") or r.get("coin") or r.get("asset", "")
        avail = (r.get("available_balance") or r.get("available")
                 or r.get("free") or r.get("balance", "0"))
        if cur:
            result.append({"source": "Wallet", "currency": cur, "available_balance": str(avail), "_raw": r})
    return result


# ─── Step 2: Balance+ Asset Summary ──────────────────────────────────────────
def fetch_asset_summary() -> list[dict]:
    data = mp_get("/flexible/api/v2/user/asset/summary")
    if data.get("code") != 0:
        raise RuntimeError(f"asset/summary error: code={data.get('code')} msg={data.get('message')}")
    inner = data.get("data") or {}
    if isinstance(inner, list):
        records = inner
    else:
        records = (inner.get("currencies") or inner.get("assets") or inner.get("list") or
                   inner.get("items") or inner.get("balances") or [])
    result = []
    for r in records:
        cur   = r.get("currency") or r.get("coin") or r.get("asset", "")
        avail = (r.get("available_balance") or r.get("available")
                 or r.get("free") or r.get("balance", "0"))
        if cur:
            result.append({"source": "Balance+", "currency": cur, "available_balance": str(avail), "_raw": r})
    return result


# ─── Step 3: Prices ───────────────────────────────────────────────────────────
def fetch_mp_rfq_price(symbol: str, qty: str) -> str | None:
    try:
        data = mp_get("/trader/v2/api/rfq-price", {"qty": qty, "side": "1", "symbol": symbol})
        if data.get("code") == 0:
            return str(data["data"]["price"])
    except Exception as e:
        log(f"  [WARN] Matrixport RFQ for {symbol}: {e}")
    return None


def fetch_coingecko_markets(ids: list[str]) -> dict:
    url = f"{COINGECKO_BASE}/coins/markets"
    params = {"vs_currency": "usd", "ids": ",".join(ids), "per_page": 20, "page": 1}
    vlog(f"GET {url}  params={params}")
    resp = requests.get(url, params=params, timeout=15)
    resp.raise_for_status()
    data = resp.json()
    vlog(f"CoinGecko response (truncated): {json.dumps(data, indent=2)[:800]}...")
    return {item["id"]: item for item in data}


def get_prices() -> dict:
    cg_data = fetch_coingecko_markets(list(COINGECKO_IDS.values()))
    prices = {}
    for token in TOKENS:
        # USDT has no RFQ pair (it's the quote currency) — use CoinGecko's live price
        if token == "USDT":
            cg_id = COINGECKO_IDS[token]
            mp_price = str(cg_data.get(cg_id, {}).get("current_price") or 1.0)
        else:
            mp_price = fetch_mp_rfq_price(*MATRIXPORT_PAIRS[token])
        cg_id    = COINGECKO_IDS[token]
        cg_item  = cg_data.get(cg_id, {})
        prices[token] = {
            "mp_price":   mp_price,
            "cg_price":   cg_item.get("current_price"),
            "cg_chg_24h": cg_item.get("price_change_percentage_24h"),
        }
    return prices


def _do_login(page, pw_expect):
    screenshot_dir = Path("Test Result/screenshots")
    screenshot_dir.mkdir(parents=True, exist_ok=True)

    for attempt in range(3):
        if attempt > 0:
            log(f"  Login retry {attempt+1}/3 — waiting 10s before retry...")
            page.wait_for_timeout(10000)

        page.goto(f"{DAM_URL}/sign-in", wait_until="domcontentloaded")
        sign_in_btn = page.locator('[data-testid="sign-in-btn"]')
        sign_in_btn.wait_for(state="visible", timeout=20000)
        page.locator('[data-testid="input-email"]').fill(DAM_EMAIL)
        page.locator('[data-testid="input-password"]').fill(DAM_PASSWORD)
        log("  Waiting for Turnstile verification to complete...")
        pw_expect(sign_in_btn).to_be_enabled(timeout=30000)
        log("  Sign In button is enabled — clicking now.")
        sign_in_btn.click()
        page.wait_for_timeout(4000)
        page.screenshot(path=str(screenshot_dir / f"login_after_click_{attempt+1}.png"), full_page=True)
        log(f"  Screenshot -> login_after_click_{attempt+1}.png  (URL: {page.url})")

        if "/sign-in" not in page.url:
            log("  Login successful.")
            page.wait_for_selector("nav, header, [class*='sidebar'], [class*='nav']", timeout=15000)
            page.wait_for_timeout(1000)
            return

        # Still on sign-in — check for error message
        err_toast = page.locator('[role="alert"].Toastify__toast--error')
        if err_toast.count() > 0:
            log(f"  [WARN] Error toast: {err_toast.inner_text()}")

    raise RuntimeError("Login failed after 3 attempts — may be rate-limited. Check login_after_click_*.png screenshots.")


# ─── Steps 5-8: Scrape DAM Portfolio ─────────────────────────────────────────
def scrape_portfolio(portfolio_id: str, headed: bool = False) -> tuple[list[dict], str | None]:
    from playwright.sync_api import sync_playwright

    rows = []
    total_value = None
    screenshot_dir = Path("Test Result/screenshots")
    screenshot_dir.mkdir(parents=True, exist_ok=True)

    from playwright.sync_api import expect as pw_expect

    auth_file = Path("dam_auth.json")

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=not headed, slow_mo=200 if headed else 0)

        # Reuse saved session if it exists (avoids repeated logins + rate limiting)
        if auth_file.exists():
            log(f"  [5] Reusing saved session from {auth_file}")
            context = browser.new_context(storage_state=str(auth_file))
            page = context.new_page()
            page.goto(f"{DAM_URL}/portfolio?portfolioId={portfolio_id}", wait_until="domcontentloaded")
            page.wait_for_timeout(3000)
            # Verify we're still logged in (not redirected to sign-in)
            if "/sign-in" in page.url:
                log("  Session expired — logging in fresh...")
                auth_file.unlink(missing_ok=True)
                page.close()
                context.close()
                context = browser.new_context()
                page = context.new_page()
                _do_login(page, pw_expect)
                context.storage_state(path=str(auth_file))
                log(f"  Session saved -> {auth_file}")
        else:
            log("  [5] Logging in to DAM (fresh session)...")
            context = browser.new_context()
            page = context.new_page()
            _do_login(page, pw_expect)
            context.storage_state(path=str(auth_file))
            log(f"  Session saved -> {auth_file}")

        log(f"  Logged in. Current URL: {page.url}")

        # Step 6: Navigate to Portfolio
        # Primary: /portfolio?portfolioId=<ID> (the format login redirects to)
        candidate_urls = [
            f"{DAM_URL}/portfolio?portfolioId={portfolio_id}",
            f"{DAM_URL}/portfolio-detail/{portfolio_id}",
            f"{DAM_URL}/portfolio/{portfolio_id}",
            f"{DAM_URL}/portfolio/{portfolio_id}/overview",
        ]
        # Check if login already landed on the correct page
        navigated = page.locator("text=David BIT Account").count() > 0
        if navigated:
            log(f"  [6] Already on portfolio page: {page.url}")

        for idx, url in enumerate(candidate_urls):
            if navigated:
                break
            log(f"  [6] Trying: {url}")
            page.goto(url, wait_until="domcontentloaded")
            page.wait_for_timeout(5000)   # allow SPA to finish rendering
            vlog(f"  Current URL after nav: {page.url}")
            vlog(f"  Page title: {page.title()}")
            page.screenshot(
                path=str(screenshot_dir / f"portfolio_attempt_{idx+1}.png"), full_page=True
            )
            if page.locator("text=David BIT Account").count() > 0:
                log(f"  Found portfolio at: {url}")
                navigated = True

        if not navigated:
            page.screenshot(path=str(screenshot_dir / "portfolio_not_found.png"), full_page=True)
            log(f"  [WARN] Could not find 'David BIT Account' on any URL. Screenshot -> portfolio_not_found.png")
            log(f"  Current URL: {page.url}")
            vlog(f"  Page text snippet: {page.locator('body').inner_text()[:800]}")
            browser.close()
            return rows, total_value

        # Step 7 & 8: Extract "David BIT Account" section via JavaScript
        # (The SPA uses complex nested divs; JS traversal is more reliable than CSS selectors)
        log("  [7] Locating 'David BIT Account' table...")
        account_heading = page.locator("text=David BIT Account").first
        try:
            account_heading.wait_for(state="visible", timeout=15000)
        except Exception:
            page.screenshot(path=str(screenshot_dir / "heading_not_visible.png"), full_page=True)
            log("  [WARN] Heading not visible. Screenshot -> heading_not_visible.png")
            browser.close()
            return rows, total_value

        page.screenshot(path=str(screenshot_dir / "portfolio_overview.png"), full_page=True)
        log("  Screenshot saved -> portfolio_overview.png")

        # Scroll the David BIT Account section into view to trigger lazy loading
        log("  Scrolling section into view and waiting for table data to load...")
        page.evaluate("""() => {
            const allEls = Array.from(document.querySelectorAll('*'));
            for (const el of allEls) {
                if (el.children.length === 0 && el.textContent.trim() === 'David BIT Account'
                        && el.className.includes('typography-title')) {
                    el.scrollIntoView({ behavior: 'smooth', block: 'center' });
                    break;
                }
            }
        }""")
        # Wait for the table rows to populate (retry up to 15s)
        for _ in range(15):
            page.wait_for_timeout(1000)
            row_count = page.evaluate("""() => {
                const allEls = Array.from(document.querySelectorAll('*'));
                for (const el of allEls) {
                    if (el.children.length === 0 && el.textContent.trim() === 'David BIT Account'
                            && el.className.includes('typography-title')) {
                        let container = el.parentElement;
                        for (let i = 0; i < 12; i++) {
                            if (!container) break;
                            if (container.querySelector('table')) {
                                return container.querySelector('table').querySelectorAll('tbody tr').length;
                            }
                            container = container.parentElement;
                        }
                    }
                }
                return 0;
            }""")
            vlog(f"  tbody tr count: {row_count}")
            if row_count > 1:  # > 1 means real data (not just "No data to display")
                break
        log(f"  Table ready with {row_count} rows.")

        log("  [8] Extracting table data via JavaScript traversal...")
        result = page.evaluate("""() => {
            // Find the heading element for "David BIT Account"
            const allText = Array.from(document.querySelectorAll('*'));
            let headingEl = null;
            for (const el of allText) {
                // Target the section heading (typography-title) not the Top Wallets widget item
                if (el.children.length === 0 && el.textContent.trim() === 'David BIT Account'
                        && el.className.includes('typography-title')) {
                    headingEl = el;
                    break;
                }
            }
            if (!headingEl) return { rows: [], totalValue: null, error: 'heading not found' };

            // Extract total value from heading's ancestor (e.g. text "David BIT Account$23.52")
            let totalValue = null;
            let ancestor = headingEl.parentElement;
            for (let i = 0; i < 6; i++) {
                if (!ancestor) break;
                const txt = (ancestor.innerText || '').replace(/\\n/g, ' ');
                const m = txt.match(/\\$([\\d,]+\\.?\\d*)/);
                if (m) { totalValue = '$' + m[1]; break; }
                ancestor = ancestor.parentElement;
            }

            // Walk up to find the container that has a table
            let container = headingEl.parentElement;
            for (let i = 0; i < 12; i++) {
                if (!container) break;
                if (container.querySelector('table')) break;
                container = container.parentElement;
            }
            if (!container || !container.querySelector('table')) {
                return { rows: [], totalValue, error: 'container with table not found' };
            }

            const table = container.querySelector('table');

            // Extract headers — first line only to avoid wrapping text
            const ths = Array.from(table.querySelectorAll('thead th'));
            const headers = ths.map(th => (th.innerText || '').split('\\n')[0].trim());

            // Extract main token rows only (skip sub-rows / account breakdown rows)
            // Main rows have Token column = a short uppercase symbol (2-10 chars, only A-Z)
            const tokenPattern = /^[A-Z]{2,10}$/;
            const trs = Array.from(table.querySelectorAll('tbody tr'));
            const rows = trs.map(tr => {
                const cells = Array.from(tr.querySelectorAll('td'));
                return cells.map(td => (td.innerText || '').split('\\n')[0].trim());
            }).filter(r => r.length >= 2 && tokenPattern.test(r[1]));

            return { headers, rows, totalValue, tr_count: trs.length };
        }""")

        vlog(f"  JS result: {json.dumps(result, indent=2)[:1000]}")

        js_headers = result.get("headers", [])
        js_rows    = result.get("rows", [])
        total_value = result.get("totalValue")

        HEADER_MAP = {
            "token":        "Token",
            "price":        "Price",
            "price (24h)":  "Price (24H)",
            "price(24h)":   "Price (24H)",
            "24h":          "Price (24H)",
            "total amount": "Total Amount",
            "amount":       "Total Amount",
            "name":         "Token",
            "value":        "Value",
            "total value":  "Total Value",
        }
        headers = [HEADER_MAP.get(h.lower(), h) for h in js_headers]
        log(f"  Headers: {js_headers} → {headers}")

        for cells in js_rows:
            vlog(f"  Row cells: {cells}")
            if cells:
                row = dict(zip(headers, cells))
                rows.append(row)

        # Total value fallback: search via Playwright if JS didn't find it
        if not total_value:
            tv_loc = page.locator("text=/Total value/i").first
            if tv_loc.count() > 0:
                total_value = tv_loc.locator("..").inner_text().strip()

        log(f"  Extracted {len(rows)} row(s).")
        log(f"  Total Value: {total_value}")

        vlog("  All extracted rows:")
        for r in rows:
            vlog(f"    {r}")

        browser.close()

    return rows, total_value


# ─── Steps 9-10: Compare ─────────────────────────────────────────────────────
def _to_float(val) -> float | None:
    if val is None:
        return None
    try:
        return float(str(val).replace(",", "").replace("$", "").replace("%", "").strip())
    except ValueError:
        return None


def within_pct(a, b, pct: float = 1.0) -> bool:
    fa, fb = _to_float(a), _to_float(b)
    if fa is None or fb is None:
        return False
    if fb == 0:
        return fa == 0
    return abs(fa - fb) / abs(fb) * 100 <= pct


def compare(api_balances: dict, prices: dict, portfolio_rows: list, port_total_raw=None):
    results = []
    calc_total = 0.0

    for row in portfolio_rows:
        raw_token   = row.get("Token", "")
        token = raw_token.upper()

        port_amount = row.get("Total Amount", "0")
        port_price  = row.get("Price (24H)", row.get("Price", "0"))
        port_value  = row.get("Value", "0")

        api_balance = api_balances.get(token, api_balances.get(raw_token.upper(), "N/A"))
        mp_price    = prices.get(token, {}).get("mp_price")
        cg_price    = prices.get(token, {}).get("cg_price")

        amount_match = "PASS" if within_pct(api_balance, port_amount) else "FAIL"
        price_match  = "PASS" if within_pct(mp_price, port_price) else "FAIL"

        fa = _to_float(port_amount) or 0.0
        fp = _to_float(mp_price) or 0.0
        calc_value  = fa * fp
        calc_total += calc_value

        value_match = "PASS" if within_pct(calc_value, port_value) else "FAIL"
        overall     = "PASS" if all(x == "PASS" for x in [amount_match, price_match, value_match]) else "FAIL"

        results.append({
            "Token":                  raw_token,
            "API Balance":            api_balance,
            "Portfolio Total Amount": port_amount,
            "Amount Match":           amount_match,
            "MP Price (USD)":         mp_price or "N/A",
            "CG Price (USD)":         cg_price or "N/A",
            "Portfolio Price":        port_price,
            "Price Match (1%)":       price_match,
            "Calc Value (amt×price)": f"{calc_value:.6f}",
            "Portfolio Value":        port_value,
            "Value Match":            value_match,
            "Overall":                overall,
        })

        vlog(
            f"  {raw_token}: api_balance={api_balance} port_amount={port_amount} amount_match={amount_match} | "
            f"mp_price={mp_price} port_price={port_price} price_match={price_match} | "
            f"calc_value={calc_value:.4f} port_value={port_value} value_match={value_match}"
        )

    # Step 10: Total Value — use the raw value scraped from the page header
    port_total = port_total_raw
    total_verdict = "PASS" if within_pct(calc_total, port_total) else "FAIL"
    vlog(f"  Total: calc={calc_total:.4f}  portfolio={port_total}  verdict={total_verdict}")

    return results, total_verdict, calc_total, port_total


# ─── Export to Excel ──────────────────────────────────────────────────────────
def export_excel(wallet, balance_plus, prices, portfolio_rows, comparison,
                 total_verdict, calc_total, port_total, out_file):
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    BLUE_FILL  = PatternFill("solid", fgColor="1E3A5F")
    WHITE_FONT = Font(color="FFFFFF", bold=True)
    PASS_FILL  = PatternFill("solid", fgColor="C6EFCE")
    FAIL_FILL  = PatternFill("solid", fgColor="FFC7CE")
    BOLD       = Font(bold=True)
    CENTER     = Alignment(horizontal="center")

    def auto_width(ws):
        for col in ws.columns:
            width = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + 4, 45)

    def write_header(ws, headers):
        ws.append(headers)
        for cell in ws[1]:
            cell.fill      = BLUE_FILL
            cell.font      = WHITE_FONT
            cell.alignment = CENTER

    wb = openpyxl.Workbook()

    # Sheet 1 — Wallet Balance
    ws1 = wb.active
    ws1.title = "Step 1 – Wallet Balance"
    write_header(ws1, ["Currency", "Available Balance"])
    for r in wallet:
        ws1.append([r["currency"], r["available_balance"]])
    auto_width(ws1)

    # Sheet 2 — Balance+
    ws2 = wb.create_sheet("Step 2 – Balance+")
    write_header(ws2, ["Currency", "Available Balance"])
    for r in balance_plus:
        ws2.append([r["currency"], r["available_balance"]])
    auto_width(ws2)

    # Sheet 3 — Prices
    ws3 = wb.create_sheet("Step 3-4 – Prices")
    write_header(ws3, ["Token", "Matrixport Price (USD)", "CoinGecko Price (USD)", "CoinGecko 24h Change (%)"])
    for token, p in prices.items():
        chg = f"{p['cg_chg_24h']:+.2f}%" if p["cg_chg_24h"] is not None else "N/A"
        ws3.append([token, p["mp_price"], p["cg_price"], chg])
    auto_width(ws3)

    # Sheet 4 — Portfolio
    ws4 = wb.create_sheet("Step 8 – Portfolio")
    if portfolio_rows:
        ph = list(portfolio_rows[0].keys())
        write_header(ws4, ph)
        for row in portfolio_rows:
            ws4.append([row.get(h, "") for h in ph])
    else:
        ws4.append(["No data scraped from portfolio"])
    auto_width(ws4)

    # Sheet 5 — Comparison
    ws5 = wb.create_sheet("Step 9-10 – Comparison")
    comp_headers = [
        "Token", "API Balance", "Portfolio Total Amount", "Amount Match",
        "MP Price (USD)", "CG Price (USD)", "Portfolio Price", "Price Match (1%)",
        "Calc Value (amt×price)", "Portfolio Value", "Value Match", "Overall",
    ]
    write_header(ws5, comp_headers)
    for i, row in enumerate(comparison, start=2):
        ws5.append([row.get(h, "") for h in comp_headers])
        fill = PASS_FILL if row.get("Overall") == "PASS" else FAIL_FILL
        for cell in ws5[i]:
            cell.fill = fill

    # Total Value summary row
    last = ws5.max_row + 2
    ws5.cell(last, 1, "TOTAL VALUE (Step 10)").font = BOLD
    ws5.cell(last, 9, f"{calc_total:.4f}").font = BOLD
    ws5.cell(last, 10, str(port_total or "N/A")).font = BOLD
    ws5.cell(last, 12, total_verdict).font = BOLD
    row_fill = PASS_FILL if total_verdict == "PASS" else FAIL_FILL
    for col in range(1, 13):
        ws5.cell(last, col).fill = row_fill

    auto_width(ws5)
    wb.save(out_file)
    log(f"  Saved -> {out_file}")


# ─── Main ─────────────────────────────────────────────────────────────────────
def main():
    global VERBOSE

    parser = argparse.ArgumentParser(description="BIT Breakdown Table Acceptance Criteria Checker")
    parser.add_argument("--portfolio-id", required=True, help="DAM Portfolio ID")
    parser.add_argument("--headed",  action="store_true", help="Show browser window")
    parser.add_argument("--verbose", action="store_true", help="Print raw API responses and row details")
    parser.add_argument("--out", default="bit_breakdown_table_report.xlsx", help="Output Excel filename")
    args = parser.parse_args()

    VERBOSE = args.verbose

    print("=" * 80)
    print("  BIT Breakdown Table — Acceptance Criteria Checker")
    print("=" * 80)
    print(f"  API Key      : {API_KEY[:20]}...")
    print(f"  Portfolio ID : {args.portfolio_id}")
    print(f"  Verbose      : {VERBOSE}")
    print()

    # ── Step 1 ─────────────────────────────────────────────────────────────────
    print("─" * 60)
    print("Step 1: Wallet Balance")
    wallet = fetch_wallet_balance()
    nonzero_wallet = [r for r in wallet if float(r["available_balance"] or 0) != 0]
    if nonzero_wallet:
        for r in nonzero_wallet:
            print(f"  {r['currency']:<10} available_balance = {r['available_balance']}")
        if len(wallet) > len(nonzero_wallet):
            print(f"  ... ({len(wallet) - len(nonzero_wallet)} zero-balance tokens hidden)")
    elif wallet:
        print(f"  All {len(wallet)} tokens have zero available_balance.")
    else:
        print("  (no records returned)")
    print()

    # ── Step 2 ─────────────────────────────────────────────────────────────────
    print("─" * 60)
    print("Step 2: Balance+ Asset Summary")
    balance_plus = fetch_asset_summary()
    nonzero_bp = [r for r in balance_plus if float(r["available_balance"] or 0) != 0]
    if nonzero_bp:
        for r in nonzero_bp:
            print(f"  {r['currency']:<10} available_balance = {r['available_balance']}")
        if len(balance_plus) > len(nonzero_bp):
            print(f"  ... ({len(balance_plus) - len(nonzero_bp)} zero-balance tokens hidden)")
    elif balance_plus:
        print(f"  All {len(balance_plus)} tokens have zero available_balance.")
    else:
        print("  (no records returned)")
    print()

    # ── Step 3 ─────────────────────────────────────────────────────────────────
    print("─" * 60)
    print("Step 3: Token Prices (Matrixport RFQ + CoinGecko 24h)")
    prices = get_prices()
    print(f"  {'Token':<8} {'MP Price':>18} {'CG Price':>18} {'24h Change':>12}")
    print(f"  {'-'*60}")
    for token, p in prices.items():
        chg = f"{p['cg_chg_24h']:+.2f}%" if p["cg_chg_24h"] is not None else "N/A"
        print(f"  {token:<8} {str(p['mp_price']):>18} {str(p['cg_price']):>18} {chg:>12}")
    print()

    # ── Steps 5-8 ──────────────────────────────────────────────────────────────
    print("─" * 60)
    print("Steps 5-8: DAM Login → Portfolio → Scrape 'David BIT Account'")
    portfolio_rows, total_value = scrape_portfolio(args.portfolio_id, headed=args.headed)
    if portfolio_rows:
        print()
        headers = list(portfolio_rows[0].keys())
        col_w = 18
        print("  " + "  ".join(h.ljust(col_w) for h in headers))
        print("  " + "-" * (col_w * len(headers) + 2 * len(headers)))
        for row in portfolio_rows:
            print("  " + "  ".join(str(row.get(h, "")).ljust(col_w) for h in headers))
        print()
        print(f"  Total Value (from page): {total_value}")
    else:
        print("  (no rows scraped)")
    print()

    # ── Steps 9-10 ─────────────────────────────────────────────────────────────
    print("─" * 60)
    print("Steps 9-10: Comparison & Calculation Verification")
    api_balances: dict[str, float] = {}
    for r in wallet + balance_plus:
        cur = r["currency"].upper()
        api_balances[cur] = api_balances.get(cur, 0.0) + float(r["available_balance"] or 0)

    comparison, total_verdict, calc_total, port_total = compare(
        api_balances, prices, portfolio_rows, port_total_raw=total_value
    )

    print()
    print(f"  {'Token':<10} {'Amt Match':<12} {'Price Match':<14} {'Value Match':<13} {'Overall'}")
    print(f"  {'-'*60}")
    for c in comparison:
        icon = "✓" if c["Overall"] == "PASS" else "✗"
        print(
            f"  {icon} {c['Token']:<9} "
            f"{c['Amount Match']:<12} {c['Price Match (1%)']:<14} "
            f"{c['Value Match']:<13} {c['Overall']}"
        )

    print()
    total_icon = "✓" if total_verdict == "PASS" else "✗"
    print(f"  {total_icon} Total Value: calculated={calc_total:.4f}  portfolio={port_total}  → {total_verdict}")
    print()

    # ── Export ─────────────────────────────────────────────────────────────────
    print("─" * 60)
    print("Step 4 & 8: Exporting Excel report...")
    export_excel(wallet, balance_plus, prices, portfolio_rows, comparison,
                 total_verdict, calc_total, port_total, args.out)

    # ── Summary ────────────────────────────────────────────────────────────────
    print()
    print("=" * 80)
    all_ok = all(c["Overall"] == "PASS" for c in comparison) and total_verdict == "PASS"
    print(f"  Final Result: {'ALL PASS ✓' if all_ok else 'SOME CHECKS FAILED ✗'}")
    print("=" * 80)


if __name__ == "__main__":
    main()
